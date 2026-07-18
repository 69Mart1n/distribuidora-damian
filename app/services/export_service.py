from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.engine import Engine

from app.database.connection import build_session_factory, session_scope
from app.database.models import ExportJob, Settings
from app.reports.receipt_pdf import build_receipt_pdf
from app.reports.wholesale_price_pdf import build_wholesale_price_pdf
from app.services.product_service import ProductService, ProductSummary
from app.services.receipt_service import ReceiptService


class ExportService:
    def __init__(self, engine: Engine, exports_dir: Path) -> None:
        self._engine = engine
        self._exports_dir = exports_dir
        self._session_factory = build_session_factory(engine)

    def export_receipt_pdf(self, receipt_id: int) -> Path:
        receipt = ReceiptService(self._engine).get_receipt(receipt_id)
        with session_scope(self._session_factory) as session:
            settings = session.get(Settings, 1)
            if settings is None:
                raise RuntimeError("No se encontró la configuración comercial.")
            session.expunge(settings)
        path = build_receipt_pdf(receipt, self._exports_dir / "boletas", settings)
        self._record("receipt", "pdf", path, 0, {}, receipt_id)
        return path

    def export_wholesale_price_pdf(
        self, *, supplier_id: int | None = None, category_id: int | None = None
    ) -> Path:
        products = self._products(supplier_id, category_id)
        path = build_wholesale_price_pdf(products, self._exports_dir / "listas_precios")
        self._record(
            "wholesale_prices",
            "pdf",
            path,
            len(products),
            {"supplier_id": supplier_id, "category_id": category_id},
        )
        return path

    def export_wholesale_price_excel(
        self, *, supplier_id: int | None = None, category_id: int | None = None
    ) -> Path:
        products = self._products(supplier_id, category_id)
        output_dir = self._exports_dir / "excel"
        output_dir.mkdir(parents=True, exist_ok=True)
        path = output_dir / f"Lista_Precios_{datetime.now():%Y-%m-%d_%H-%M-%S}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Precios"
        sheet.append(["Código", "Producto", "Presentación", "Proveedor", "Categoría", "Precio"])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="34533C")
        for product in products:
            sheet.append(self._row(product))
        widths = [14, 42, 24, 24, 22, 14]
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[chr(64 + index)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(path)
        self._record(
            "wholesale_prices",
            "xlsx",
            path,
            len(products),
            {"supplier_id": supplier_id, "category_id": category_id},
        )
        return path

    def export_wholesale_price_csv(
        self, *, supplier_id: int | None = None, category_id: int | None = None
    ) -> Path:
        products = self._products(supplier_id, category_id)
        output_dir = self._exports_dir / "csv"
        output_dir.mkdir(parents=True, exist_ok=True)
        path = output_dir / f"Lista_Precios_{datetime.now():%Y-%m-%d_%H-%M-%S}.csv"
        with path.open("w", encoding="utf-8-sig", newline="") as destination:
            writer = csv.writer(destination, delimiter=";")
            writer.writerow(
                ["Código", "Producto", "Presentación", "Proveedor", "Categoría", "Precio"]
            )
            for product in products:
                writer.writerow(self._row(product))
        self._record(
            "wholesale_prices",
            "csv",
            path,
            len(products),
            {"supplier_id": supplier_id, "category_id": category_id},
        )
        return path

    def _products(self, supplier_id: int | None, category_id: int | None) -> list[ProductSummary]:
        return ProductService(self._engine).list_products(
            supplier_id=supplier_id, category_id=category_id
        )

    @staticmethod
    def _row(product: ProductSummary) -> list[object]:
        return [
            product.code,
            product.name,
            product.presentation,
            product.supplier,
            product.category,
            int(product.wholesale_price) if product.wholesale_price is not None else "",
        ]

    def _record(
        self,
        export_type: str,
        file_format: str,
        path: Path,
        total_products: int,
        filters: dict[str, int | None],
        receipt_id: int | None = None,
    ) -> None:
        with session_scope(self._session_factory) as session:
            session.add(
                ExportJob(
                    export_type=export_type,
                    format=file_format,
                    filename=path.name,
                    file_path=str(path.resolve()),
                    filters_json=json.dumps(filters),
                    total_products=total_products or None,
                    receipt_id=receipt_id,
                )
            )
