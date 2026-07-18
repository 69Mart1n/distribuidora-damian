from __future__ import annotations

import csv
import re
import shutil
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.engine import Engine
from sqlalchemy.orm import Session

from app.database.connection import build_session_factory, session_scope
from app.database.models import Category, ImportJob, ImportRow, Product, Supplier
from app.utils.money import require_whole_number, to_money
from app.utils.text import normalize_text

DEFAULT_PRICE_LIST_PDF = Path(r"C:\Users\56944438\Downloads\Precios Por Mayor.pdf")

SUPPLIER_HEADERS_TO_IGNORE = {
    "",
    "producto kilos precio",
    "producto presentacion precio",
    "producto presentación precio",
    "producto tamaño precio",
    "producto kilos presentacion precio",
    "producto kilos presentación precio",
    "producto kilos descuento",
}


@dataclass(frozen=True)
class ImportResult:
    detected: int
    created: int
    updated: int
    skipped: int
    source: Path


@dataclass(frozen=True)
class ParsedProductRow:
    supplier: str
    name: str
    presentation: str
    price: Decimal | None
    raw_line: str


@dataclass
class PreviewImportRow:
    row_number: int
    supplier: str
    name: str
    presentation: str
    price: Decimal | None
    action: str
    linked_product_id: int | None = None
    selected: bool = True
    error: str = ""


@dataclass(frozen=True)
class ImportPreview:
    source: Path
    rows: list[PreviewImportRow]


class PriceListImportService:
    def __init__(self, engine: Engine, storage_dir: Path | None = None) -> None:
        self._session_factory = build_session_factory(engine)
        self._storage_dir = storage_dir

    def preview_file(self, source: Path) -> ImportPreview:
        suffix = source.suffix.lower()
        if suffix == ".pdf":
            parsed = parse_price_list_pdf(source)
        elif suffix in {".xlsx", ".xlsm"}:
            parsed = _parse_spreadsheet(source)
        elif suffix == ".csv":
            parsed = _parse_csv(source)
        else:
            raise ValueError("El formato debe ser PDF, XLSX o CSV.")
        preview_rows: list[PreviewImportRow] = []
        with session_scope(self._session_factory) as session:
            for number, row in enumerate(parsed, 1):
                error = ""
                if not row.name or not row.supplier:
                    error = "Falta producto o proveedor."
                elif row.price is not None:
                    try:
                        require_whole_number(row.price, "El precio")
                    except ValueError as exception:
                        error = str(exception)
                supplier = session.scalar(
                    select(Supplier).where(Supplier.normalized_name == normalize_text(row.supplier))
                )
                product = (
                    _find_product(session, supplier.id, row.name, row.presentation)
                    if supplier
                    else None
                )
                action = "invalid" if error else "update" if product else "create"
                preview_rows.append(
                    PreviewImportRow(
                        number,
                        row.supplier,
                        row.name,
                        row.presentation,
                        row.price,
                        action,
                        product.id if product else None,
                        not bool(error),
                        error,
                    )
                )
        return ImportPreview(source, preview_rows)

    def commit_preview(self, preview: ImportPreview) -> ImportResult:
        if not preview.source.exists():
            raise FileNotFoundError("No se encontró el archivo de importación.")
        stored = preview.source
        if self._storage_dir:
            self._storage_dir.mkdir(parents=True, exist_ok=True)
            stored = self._storage_dir / preview.source.name
            if stored.resolve() != preview.source.resolve():
                shutil.copy2(preview.source, stored)
        with session_scope(self._session_factory) as session:
            job = ImportJob(
                original_filename=preview.source.name,
                stored_filename=stored.name,
                file_type=preview.source.suffix.lower().lstrip("."),
                status="processing",
                total_rows_detected=len(preview.rows),
            )
            session.add(job)
            session.flush()
            created = updated = skipped = errors = 0
            next_code_number = _next_product_number(session)
            for row in preview.rows:
                selected = row.selected and row.action != "invalid"
                linked_id = row.linked_product_id
                if selected:
                    supplier = _get_or_create_supplier(session, row.supplier)
                    category = _get_or_create_category(
                        session,
                        infer_category(
                            ParsedProductRow(
                                row.supplier,
                                row.name,
                                row.presentation,
                                row.price,
                                "",
                            )
                        ),
                    )
                    product = session.get(Product, linked_id) if linked_id else None
                    if product is None:
                        product = Product(
                            internal_code=f"P{next_code_number:04d}",
                            name=row.name.strip(),
                            normalized_name=normalize_text(row.name),
                            supplier_id=supplier.id,
                            category_id=category.id,
                            presentation=_normalize_presentation(row.presentation),
                            wholesale_price=to_money(row.price) if row.price is not None else None,
                            requires_review=row.price is None,
                            notes=f"Importado desde {preview.source.name}",
                        )
                        session.add(product)
                        session.flush()
                        next_code_number += 1
                        created += 1
                    else:
                        # Las coincidencias exactas actualizan; precios manuales solo cambian
                        # cuando la fila importada trae un valor válido.
                        product.category_id = category.id
                        if row.price is not None:
                            product.wholesale_price = to_money(row.price)
                            product.requires_review = False
                        updated += 1
                    linked_id = product.id
                else:
                    skipped += 1
                    errors += int(bool(row.error))
                session.add(
                    ImportRow(
                        import_id=job.id,
                        row_number=row.row_number,
                        supplier_detected=row.supplier,
                        product_name_detected=row.name,
                        presentation_detected=row.presentation,
                        price_detected=row.price,
                        status="imported" if selected else "skipped",
                        error_message=row.error or None,
                        linked_product_id=linked_id,
                    )
                )
            job.status = "completed"
            job.total_rows_imported = created + updated
            job.total_rows_skipped = skipped
            job.total_errors = errors
            return ImportResult(len(preview.rows), created, updated, skipped, preview.source)

    def import_pdf(self, pdf_path: Path = DEFAULT_PRICE_LIST_PDF) -> ImportResult:
        rows = parse_price_list_pdf(pdf_path)
        with session_scope(self._session_factory) as session:
            created = 0
            updated = 0
            skipped = 0
            next_code_number = _next_product_number(session)
            for row in rows:
                if not row.name:
                    skipped += 1
                    continue
                supplier = _get_or_create_supplier(session, row.supplier)
                category = _get_or_create_category(session, infer_category(row))
                product = _find_product(session, supplier.id, row.name, row.presentation)
                if product is None:
                    product = Product(
                        internal_code=f"P{next_code_number:04d}",
                        name=row.name,
                        normalized_name=normalize_text(row.name),
                        supplier_id=supplier.id,
                        category_id=category.id,
                        presentation=row.presentation,
                        wholesale_price=to_money(row.price) if row.price is not None else None,
                        requires_review=row.price is None,
                        notes=f"Importado desde {pdf_path.name}",
                    )
                    session.add(product)
                    created += 1
                    next_code_number += 1
                else:
                    product.category_id = category.id
                    product.wholesale_price = to_money(row.price) if row.price is not None else None
                    product.requires_review = row.price is None
                    updated += 1
            return ImportResult(
                detected=len(rows),
                created=created,
                updated=updated,
                skipped=skipped,
                source=pdf_path,
            )


def parse_price_list_pdf(pdf_path: Path) -> list[ParsedProductRow]:
    if not pdf_path.exists():
        return []

    rows: list[ParsedProductRow] = []
    current_supplier = ""
    with pdfplumber.open(pdf_path) as document:
        for page in document.pages:
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            for raw_line in text.splitlines():
                line = _clean_line(raw_line)
                normalized = normalize_text(line)
                if normalized.startswith("--- page"):
                    continue
                if normalized in SUPPLIER_HEADERS_TO_IGNORE or normalized.startswith("producto "):
                    continue
                if "$" not in line:
                    if _looks_like_supplier(line):
                        current_supplier = line
                    continue
                if not current_supplier:
                    current_supplier = "Otros"
                parsed = _parse_product_line(current_supplier, line)
                if parsed is not None:
                    rows.append(parsed)
    return rows


def import_initial_pdf_if_available(engine: Engine) -> ImportResult | None:
    if not DEFAULT_PRICE_LIST_PDF.exists():
        return None
    session_factory = build_session_factory(engine)
    with session_scope(session_factory) as session:
        product_count = session.scalar(select(func.count(Product.id))) or 0
    if product_count >= 100:
        return None
    return PriceListImportService(engine).import_pdf(DEFAULT_PRICE_LIST_PDF)


def infer_category(row: ParsedProductRow) -> str:
    text = normalize_text(f"{row.supplier} {row.name} {row.presentation}")
    if any(word in text for word in ["collar", "oreja", "hueso", "snak", "snack", "palito"]):
        return "Accesorios"
    if any(word in text for word in ["nexgard", "simparica", "pipeta", "antiparasitario"]):
        return "Antiparasitarios"
    if any(
        word in text for word in ["semilla", "maiz", "arroz", "alpiste", "polenta", "afrechillo"]
    ):
        return "Semillas"
    if any(word in text for word in ["cerdo", "ovino", "equino", "conejo", "ponedora"]):
        return "Animales de granja"
    if any(word in text for word in ["gato", "cat", "whiskas", "felix", "gatito"]):
        return "Gatos"
    if any(word in text for word in ["perro", "dog", "can", "cachorro", "adulto"]):
        return "Perros"
    return "Otros"


def _parse_product_line(supplier: str, line: str) -> ParsedProductRow | None:
    match = re.search(r"\$\s*(?P<price>[\d.,]*)\s*$", line)
    if not match:
        return None
    price_text = match.group("price").strip()
    price = _parse_price(price_text)
    body = line[: match.start()].strip(" -")
    if not body:
        return None
    name, presentation = _split_name_presentation(body)
    return ParsedProductRow(
        supplier=supplier,
        name=name.strip(),
        presentation=_normalize_presentation(presentation.strip() or "Unidad"),
        price=price,
        raw_line=line,
    )


def _split_name_presentation(body: str) -> tuple[str, str]:
    lower = body.lower()
    if "collar " in lower and not re.search(r"\d", body):
        return body, "Unidad"

    special = re.match(
        r"(?P<name>.+?)\s+(?P<presentation>\d+(?:[,.]\d+)?\s*a\s*"
        r"\d+(?:[,.]\d+)?\s*kg\s+.+)$",
        body,
        flags=re.IGNORECASE,
    )
    if special:
        return special.group("name"), special.group("presentation")

    suffix_patterns = [
        r"\d+(?:[,.]\d+)?\s*g\s+caja\s+\d+\s*u",
        r"\d+(?:[,.]\d+)?\s*kg\s+caja\s+\d+\s*u",
        r"\d+(?:[,.]\d+)?\s*kg\s*\+\s*\d+(?:[,.]\d+)?\s*kg(?:\s+gratis)?",
        r"\d+(?:[,.]\d+)?\s*\+\s*\d+(?:[,.]\d+)?\s*kg",
        r"\d+(?:[,.]\d+)?\s*kg(?:\s+gratis)?",
        r"\d+(?:[,.]\d+)?\s*k(?:\s+kg)?",
        r"\d+x\d+\s*g",
        r"\d+(?:[,.]\d+)?\s*g",
        r"\d+-\d+\s*x\d+\s+unidades",
        r"\d+\s+comprimidos",
        r"\d+\s+comp\.\s+en\s+sobre",
        r"\d+\+\d+",
    ]
    for pattern in suffix_patterns:
        match = re.search(rf"(?P<presentation>{pattern})$", body, flags=re.IGNORECASE)
        if match:
            name = body[: match.start()].strip(" -")
            return name or body, match.group("presentation")
    return body, "Unidad"


def _normalize_presentation(value: str) -> str:
    value = value.replace("—", "").strip()
    value = re.sub(r"(?i)(\d)(kg|k|g)\b", r"\1 \2", value)
    value = re.sub(r"(?i)\bk\b", "kg", value)
    value = re.sub(r"\s+", " ", value)
    return value


def _parse_price(value: str) -> Decimal | None:
    if not value:
        return None
    normalized = value.replace(".", "").replace(",", ".")
    return Decimal(normalized)


def _clean_line(line: str) -> str:
    return re.sub(r"\s+", " ", line).strip()


def _looks_like_supplier(line: str) -> bool:
    normalized = normalize_text(line)
    if normalized in SUPPLIER_HEADERS_TO_IGNORE:
        return False
    if re.search(r"\d", line):
        return False
    return len(line) <= 40


def _get_or_create_supplier(session: Session, name: str) -> Supplier:
    normalized = normalize_text(name)
    supplier = session.scalar(select(Supplier).where(Supplier.normalized_name == normalized))
    if supplier is None:
        supplier = Supplier(name=name, normalized_name=normalized)
        session.add(supplier)
        session.flush()
    return supplier


def _get_or_create_category(session: Session, name: str) -> Category:
    normalized = normalize_text(name)
    category = session.scalar(select(Category).where(Category.normalized_name == normalized))
    if category is None:
        category = Category(name=name, normalized_name=normalized)
        session.add(category)
        session.flush()
    return category


def _find_product(
    session: Session,
    supplier_id: int,
    name: str,
    presentation: str,
) -> Product | None:
    return session.scalar(
        select(Product).where(
            Product.supplier_id == supplier_id,
            Product.normalized_name == normalize_text(name),
            Product.presentation == _normalize_presentation(presentation),
        )
    )


def _next_product_number(session: Session) -> int:
    last_id = session.scalar(select(Product.id).order_by(Product.id.desc()).limit(1)) or 0
    return last_id + 1


def _parse_spreadsheet(path: Path) -> list[ParsedProductRow]:
    if not path.exists():
        raise FileNotFoundError("No se encontró el archivo seleccionado.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    values = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    return _parse_tabular_rows(values)


def _parse_csv(path: Path) -> list[ParsedProductRow]:
    if not path.exists():
        raise FileNotFoundError("No se encontró el archivo seleccionado.")
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        sample = source.read(4096)
        source.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        return _parse_tabular_rows(list(csv.reader(source, dialect)))


def _parse_tabular_rows(values: list[tuple[object, ...] | list[str]]) -> list[ParsedProductRow]:
    if not values:
        return []
    headers = [normalize_text(str(value or "")) for value in values[0]]

    def column(*names: str) -> int | None:
        for index, header in enumerate(headers):
            if any(name in header for name in names):
                return index
        return None

    supplier_column = column("proveedor", "marca")
    name_column = column("producto", "nombre", "descripcion")
    presentation_column = column("presentacion", "tamano", "kilos", "unidad")
    price_column = column("precio", "mayorista", "importe")
    if name_column is None or price_column is None:
        raise ValueError("No se encontraron las columnas Producto y Precio.")
    rows: list[ParsedProductRow] = []
    for values_row in values[1:]:
        row = list(values_row)

        def cell(index: int | None, default: str = "", current_row=row) -> str:
            if index is None or index >= len(current_row) or current_row[index] is None:
                return default
            return str(current_row[index]).strip()

        name = cell(name_column)
        if not name:
            continue
        raw_price = cell(price_column).replace("$", "").strip()
        try:
            original_price = row[price_column] if price_column < len(row) else None
            if isinstance(original_price, int | float | Decimal):
                price = Decimal(str(original_price))
            else:
                price = _parse_price(raw_price)
        except Exception:
            price = None
        rows.append(
            ParsedProductRow(
                supplier=cell(supplier_column, "Otros"),
                name=name,
                presentation=_normalize_presentation(cell(presentation_column, "Unidad")),
                price=price,
                raw_line=" | ".join(str(value or "") for value in row),
            )
        )
    return rows
